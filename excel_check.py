# Otworzenie EXCELa, otworzenie pierwszego arkusza;
# - znalezienie max ilości wierszy
# - znalezienie kolumn: part number, rysunek, i obróbek
# - zwrócenie wszystkich tych wartości
import openpyxl


def excel_check(bom_path):
    wb = openpyxl.load_workbook(bom_path)
    type(wb)
    arkusze = wb.sheetnames
    sheet = wb[arkusze[0]]

    max_column = sheet.max_column
    max_row = sheet.max_row

    if "BOM" in sheet.cell(row=1, column=1).value.upper():
        row_to_check = 2
    else:
        row_to_check = 1


    for i in range(1, max_column + 1):

        value = sheet.cell(row=row_to_check, column=i).value
        if "TCH" in value.upper() and "1" in value.upper():
            kolumna_tch1 = i

        if "TCH" in value.upper() and "2" in value.upper():
            kolumna_tch2 = i

        if "TCH" in value.upper() and "3" in value.upper():
            kolumna_tch3 = i

        if "RYSUNEK" in value.upper():
            kolumna_rysunek = i

        if "PART" in value.upper() and "NUMBER" in value.upper():
            kolumna_part_number = i

        if "ITEM" in value.upper() and "NO" in value.upper():
            kolumna_item_number = i

        if "QTY" in value.upper() and "TOTAL" not in value.upper():
            kolumna_qty = i

        if "QTY" in value.upper() and "TOTAL" in value.upper():
            kolumna_qty_total = i

    # print(kolumna_item_number)
    # print(kolumna_part_number)
    # print(kolumna_qty)
    # print(kolumna_qty_total)
    # print(kolumna_tch1)
    # print(kolumna_tch2)
    # print(kolumna_tch3)
    # print(kolumna_rysunek)

    return (kolumna_item_number, kolumna_part_number, kolumna_qty, kolumna_qty_total, kolumna_tch1, kolumna_tch2,
            kolumna_tch3, kolumna_rysunek, max_row)
