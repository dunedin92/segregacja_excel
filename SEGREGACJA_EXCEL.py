#!/usr/bin/env python3
# coding: utf-8
# -*- coding: utf-8 -*-

import linecache
import shutil
import os
import openpyxl

from finding_bom import finding_bom
from excel_check import excel

## ================================================================================================
## Funkcja przeszukująca excel i tworząca fodlery zgodnie z kolumnami tch1 tch2 tch3 rysunek
def folder_creation(source, destination, kolumna_part_number, kolumna_tch1, kolumna_tch2, kolumna_tch3, kolumna_rysunek, max_row):

  formats = [".pdf", ".dxf", ".step", ".stl"]
  
  for i in range(2, max_row+1):

    part_number = sheet.cell(row=i, column=kolumna_part_number).value
    rysunek = sheet.cell(row=i, column=kolumna_rysunek).value
    tch1 = sheet.cell(row=i, column=kolumna_tch1).value
    tch2 = sheet.cell(row=i, column=kolumna_tch2).value
    tch3 = sheet.cell(row=i, column=kolumna_tch3).value
    
    if "WYKONANY" in rysunek.upper() or ("RYSUNEK" in rysunek.upper() and "SPAWALNICZY" in rysunek.upper()):
      if tch1 != "-":
        if tch2 != "-":
          if tch3 != "-":
            folder_name = tch1 + "-" + tch2 + "-" + tch3
          else:
            folder_name = tch1 + "-" + tch2
        else:
          folder_name = tch1

        folder_destination = os.path.join(destination, folder_name)
        
        if os.path.exist(folder_destiation):
          for name in formats:
            part = part_number + name
            part_destination = os.path.join(destination, part)

            for path, dirs, files in os.walk(source):
              part_source = os.path.join(path, part)
              if os.path.exist(part_source):
                if os.path.exist(part_destination):
                  print("taki plik został już wczesniej przeniesiony")
                  break
                else:
                  shutil.copy(part_source, part_destination)
                  break
              
        
      else:
        print(part_number)
        print("dla tego pliku nie ma przypisanej obróbki")
        ## dopisać wrzucanie tej informacji do pliku tekstowego z brakującymi plikami

      
  return(0)

## ------------------------------------------------------------------------------------------------

## ================================================================================================
## Funkcja przegladajaca excel i przenoszaca pliki do odpowiednich folderów
## W zależności od folderów kopiuje odpowiednie pliki (do C pdf i dxf, do F P pdf i step, do DRUK 3D stl, itp)
## Tworzy plik tekstowy z brakującymi pliki i dopisuje do jakiego folderu powinien trafic brakujacy plik

def segregation():
  return()

## ------------------------------------------------------------------------------------------------




## ================================================================================================
## ================================================================================================
##
##                              ======>> PĘTLA GŁÓWNA PROGRAMU <<======
##
## ================================================================================================
## ================================================================================================

#sciezka której kopiujemy pliki:
source = "D:\PROGRAMOWANIE\segregation_files"

##scieżka do BOM'u i miejsce gdzie będą wrzucone posegregowane pliki:
##destination = "D:\PROGRAMOWANIE\pliki"

#pobranie linku do foledru gdzie jest bom i beda kopiowane pliki:
destination = input("Podaj scieżkę do pliku BOM, tutaj zostaną skopiowane posegregowane pliki.\n"
                    " Upewnij się, że w folderze jest plik .xlsx i jest to BOM.\n ==>")
  
if finding_bom(destination):
  bom_path = finding_bom(destination)
  print("=" * 60)
  print(" Znaleziono plik z BOMem:\n"
        " Dokładna ścieżka do pliku z BOMem to:\n")
  print("====>   " + bom_path + "   <====")
  print("\n" + "=" * 60 +"\n")
  
else:
  print("nie znaleziono pliku excel z BOMem")
  exit()
  
part_number, tch1, tch2, tch3, rysunek, max_row = excel(bom_path)
