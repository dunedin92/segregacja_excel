#!/usr/bin/env python3
# coding: utf-8
# -*- coding: utf-8 -*-

import linecache
import shutil
import os
import openpyxl

def qty_total_calculation(bom_path):

      ##otworzenie arkusza BOM'u
      wb = openpyxl.load_workbook(bom_path)
      type(wb)

      arkusze = wb.get_sheet_names()

      sheet =wb[arkusze[1]]
      sheet = wb.active
      print (sheet)

      max_column = sheet.max_column
      max_row = sheet.max_row
      print(max_column)
      print(max_row)



      ##pobranie lokalizacji kolumn z nazwa i rodzajem obróbki
      for i in range(1,max_column+1):
            ##print(sheet.cell(row=1, column=i).value)
            if sheet.cell(row=1, column=i).value == "ITEM NO.":
                  kolumna_item_number = i
            if sheet.cell(row=1, column=i).value == "QTY.":
                  kolumna_qty = i
            if sheet.cell(row=1, column=i).value == "QTY. TOTAL":
                  kolumna_qty_total = i

      print(kolumna_item_number)
      print(kolumna_qty)
      print(kolumna_qty_total)

      ##liczenie ilosci kropek i przecinkow w numerze
      for i in range(2,max_row+1):
            cell_value_actual = sheet.cell(row=i, column=kolumna_item_number).value
            cell_value_last = sheet.cell(row=i-1, column=kolumna_item_number).value
            print(cell_value_actual)
            cell_value_actual = str(cell_value_actual)
            cell_value_last = str(cell_value_last)
            qt_dots_actual=(cell_value_actual.count("."))+(cell_value_actual.count(","))
            qt_dots_last=(cell_value_last.count("."))+(cell_value_last.count(","))
            #print(qt_dots_actual)
            #print('roznica w zagniezdzeniu:')
            #print(qt_dots_actual - qt_dots_last)

            if (qt_dots_actual==0):
                  qty_i=sheet.cell(row=i, column=kolumna_qty).value
                  sheet.cell(row=i, column=kolumna_qty_total).value = qty_i
                  #print('wpisujemy wartosc qty_total:')
                  #print(sheet.cell(row=i, column=kolumna_qty_total).value)
            elif qt_dots_actual - qt_dots_last==1:
                  qty_i_last=sheet.cell(row=i-1, column=kolumna_qty_total).value
                  qty_i=sheet.cell(row=i, column=kolumna_qty).value
                  sheet.cell(row=i, column=kolumna_qty_total).value = qty_i_last * qty_i
                  #print('wpisujemy wartosc qty_total:')
                  #print(sheet.cell(row=i, column=kolumna_qty_total).value)
            elif qt_dots_actual - qt_dots_last==0 or qt_dots_actual - qt_dots_last==-1:
                  #print('ten sam poziom zagniezdzenia co wyzej')
                  for j in range(i-1, 0, -1):
                        temporary_cell_value = sheet.cell(row=j, column=kolumna_item_number).value
                        temporary_cell_value = str(temporary_cell_value)
                        qt_dots_temporary = (temporary_cell_value.count("."))+(temporary_cell_value.count(","))
                        if qt_dots_actual - qt_dots_temporary == 1:
                              #print('znaleziono nizszy poziom zagniezdzenia')
                              qty_i=sheet.cell(row=i, column=kolumna_qty).value
                              qty_i_temporary = sheet.cell(row=j, column=kolumna_qty_total).value
                              sheet.cell(row=i, column=kolumna_qty_total).value = qty_i * qty_i_temporary
                              #print('wpisujemy wartosc qty_total:')
                              #print(sheet.cell(row=i, column=kolumna_qty_total).value)
                              break


      wb.save(file_list)
      return True