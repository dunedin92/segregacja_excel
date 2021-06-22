# Funkcja przeszukująca excel i tworząca fodlery zgodnie z kolumnami tch1 tch2 tch3 rysunek
# Przegladajaca excel i przenoszaca pliki do odpowiednich folderów
# W zależności od folderów kopiuje odpowiednie pliki (do C pdf i dxf, do F P pdf i step, do DRUK 3D stl, itp)
# Tworzy plik tekstowy z brakującymi pliki i dopisuje do jakiego folderu powinien trafic brakujacy plik
import os
import shutil
import openpyxl
import win32pipe
import win32file
import subprocess

def file_segregation(source, destination, bom_path,  kolumna_part_number, kolumna_tch1, kolumna_tch2, kolumna_tch3,
                     kolumna_rysunek, max_row):
    wb = openpyxl.load_workbook(bom_path)
    type(wb)
    arkusze = wb.sheetnames
    sheet = wb[arkusze[0]]

    formats = [".dxf", ".step"]
    no_file_in_surce = []

    class PipeServer:
        def __init__(self, pipeName):
            self.pipe = win32pipe.CreateNamedPipe(r'\\.\pipe\\' + pipeName, win32pipe.PIPE_ACCESS_OUTBOUND,
                                                  win32pipe.PIPE_TYPE_MESSAGE | win32pipe.PIPE_READMODE_MESSAGE | win32pipe.PIPE_WAIT,
                                                  1, 65536, 65536, 1, None)

        # Carefull, this blocks until a connection is established
        def connect(self):
            win32pipe.ConnectNamedPipe(self.pipe, None)

        # Message without tailing '\n'
        def write(self, message):
            win32file.WriteFile(self.pipe, message.encode() + b'\n')

        def close(self):
            win32file.CloseHandle(self.pipe)

        def read(self):
            code, data = win32pipe.ReadFile(self.pipe, 2048)
            if code == 0:
                return data
            else:
                print('Saw result code of %d with data len=%d' % (code, len(data)))

    subprocess.Popen("apptest.exe")
    print("otworzono plik c#")
    t = PipeServer("CSServer")

    for i in range(2, max_row + 1):

        part_number = sheet.cell(row=i, column=kolumna_part_number).value
        part_number = part_number.lstrip()
        print(part_number)

        rysunek = sheet.cell(row=i, column=kolumna_rysunek).value
        tch1 = sheet.cell(row=i, column=kolumna_tch1).value
        tch2 = sheet.cell(row=i, column=kolumna_tch2).value
        tch3 = sheet.cell(row=i, column=kolumna_tch3).value

# wyszukanie plików w BOM ktore trzeba przekonwertować i przenieść
        if "WYKONANY" in rysunek.upper() or ("RYSUNEK" in rysunek.upper() and "SPAWALNICZY" in rysunek.upper()):
            print("plik do konwersji")
            if tch1 != "-":
                if tch2 != "-":
                    if tch3 != "-":
                        folder_name = tch1.upper() + "+" + tch2.upper() + "+" + tch3.upper()
                    else:
                        folder_name = tch1.upper() + "+" + tch2.upper()
                else:
                    folder_name = tch1.upper()

                folder_destination = os.path.join(destination, folder_name)

                if not os.path.exists(folder_destination):
                    os.mkdir(folder_destination)
                    print("folder docelowy: " + folder_destination)
            else:
                no_file_in_surce.append(part_number + " - brak podanej obróbki w BOM")

# znalezienie scieżki do pliku modelu solida
            for path, dirs, files in os.walk(source):
                print("szukana sciezka: " + path)

                part_number_sldprt = part_number + ".sldprt"
                print("numer czesci solidworks: " + part_number_sldprt)
                part_source = os.path.join(path, part_number_sldprt)
                print("sciezka zrodłowa pliku: " +part_source)
                part_destination = os.path.join(folder_destination, part_number)

                part_destination = part_destination + ".pdf"
                print("sciezka docelowa pliku: " + part_destination)

                if os.path.exists(part_source):
                    print("sciezka do pliku sldprt istnieje")

                    if os.path.exists(part_destination):
                        print("plik został już przekonwertowany i przeniesiony do odpowiedniego miejsca")
                    else:
                        print("Python: wysyłamy sciezke do pliku do konwersji na inne formaty:")

                        print("stworzenie pipeservera")
                        t.connect()

                        if t.read() == 'potwierdzenie połączenia':
                            print("potwierdozno połaczenie serwerów")
                        print("połączenie z pipe serwerem")
                        part_source = part_source[0:part_source.rfind(".")]
                        t.write(part_source)
                        print("wysyłamy do C#:  " + part_source)

                        if t.read() != "OK":
                            print("błąd w trakcie zapisu do innego formatu")
                            no_file_in_surce.append(part_number + " - " + folder_name)
                        else:
                            print('udało sie wslać sciezke do pliku')
                            print("przenosimy przekonwertowane pliki do odpowiedniego folderu")

                            for format in formats:
                                if os.path.exists(part_source + format):
                                    shutil.move(part_source, folder_destination)
                                else:
                                    print("pliku o danym formacie nie ma w folderze")
                                    for i in no_file_in_surce:
                                        if part_number in i:
                                            print("brak pliku został juz odnotowany")
                                        else:
                                            no_file_in_surce.append(part_number + " - " + folder_name)

    return no_file_in_surce
