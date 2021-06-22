import os
import openpyxl
import time

def temp_file_list(source, destination, bom_path, kolumna_part_number, kolumna_tch1, kolumna_tch2, kolumna_tch3,
                   kolumna_rysunek, max_row):
    wb = openpyxl.load_workbook(bom_path)
    type(wb)
    arkusze = wb.sheetnames
    sheet = wb[arkusze[0]]

    no_file_in_surce = []
    temp_file_txt = []

    # przeglądanie BOMu linijka po linijce
    for i in range(2, max_row + 1):

        part_number = sheet.cell(row=i, column=kolumna_part_number).value
        part_number = part_number.lstrip()
        part_number_sldprt = part_number + ".sldprt"
        part_number_sldasm = part_number + ".sldasm"

        rysunek = sheet.cell(row=i, column=kolumna_rysunek).value
        tch1 = sheet.cell(row=i, column=kolumna_tch1).value
        tch2 = sheet.cell(row=i, column=kolumna_tch2).value
        tch3 = sheet.cell(row=i, column=kolumna_tch3).value

        # wyszukanie plików w BOM ktore trzeba przekonwertować i przenieść
        if "WYKONANY" in rysunek.upper() or ("RYSUNEK" in rysunek.upper() and "SPAWALNICZY" in rysunek.upper()):
            print("plik do konwersji")
            if tch1 != "-":
                if tch2 != "-":
                    if tch3 != "-":
                        folder_name = tch1.upper() + "+" + tch2.upper() + "+" + tch3.upper()
                    else:
                        folder_name = tch1.upper() + "+" + tch2.upper()
                else:
                    folder_name = tch1.upper()

                folder_destination = os.path.join(destination, folder_name)

                if not os.path.exists(folder_destination):
                    os.mkdir(folder_destination)
                    print("folder docelowy: " + folder_destination)

                # znajdowanie sciezki do pliku znalezionego w BOM
                for path, dirs, files in os.walk(source):

                    file_location = os.path.join(path, part_number_sldprt)
                    file_location_asm = os.path.join(path, part_number_sldasm)
                    print("szukana sciezka: " + file_location + "  lub: " + file_location_asm)

                    if os.path.exists(file_location) or os.path.exists(file_location_asm):
                        print(" => podana ścieżka istnieje")

                        line = os.path.join(path, part_number) + " => " + folder_name
                        print(line)

                        if line in temp_file_txt:
                            print("element został już dodany do listy")
                        else:
                            temp_file_txt.append(line)
                            print("dodano do listy")

            else:
                no_file_in_surce.append(part_number + " - brak podanej obróbki w BOM")




    plik = open("temp_file_txt.txt", "w", encoding='utf8')

    for name in temp_file_txt:
        plik.write(name)
        plik.write("\n")
    plik.close()

    modification_time = time.ctime(os.path.getmtime("temp_file_txt.txt"))

    return no_file_in_surce, modification_time
